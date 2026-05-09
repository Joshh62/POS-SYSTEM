from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas
from app.dependencies import get_current_user, require_role, SUPERADMIN_ROLE
from datetime import date, datetime
import openpyxl
import csv
import io
import pytz

router = APIRouter(prefix="/products", tags=["Products"])

LAGOS = pytz.timezone("Africa/Lagos")

def now_lagos():
    return datetime.now(LAGOS).replace(tzinfo=None)


def write_audit(db, user_id, action, table, record_id, description):
    try:
        db.add(models.AuditLog(
            user_id=user_id, action=action, table_name=table,
            record_id=record_id, description=description,
        ))
    except Exception as e:
        print(f"[Audit] {e}")


def _get_business_branches(db, user):
    if user.role == SUPERADMIN_ROLE:
        return db.query(models.Branch).all()
    return db.query(models.Branch).filter(
        models.Branch.business_id == user.business_id
    ).all()


def _product_dict(p, db):
    supplier = None
    if p.supplier_id:
        s = db.query(models.Supplier).filter(
            models.Supplier.supplier_id == p.supplier_id
        ).first()
        if s:
            supplier = {"supplier_id": s.supplier_id, "supplier_name": s.supplier_name}
    return {
        "product_id":    p.product_id,
        "business_id":   p.business_id,
        "product_name":  p.product_name,
        "barcode":       p.barcode,
        "category_id":   p.category_id,
        "cost_price":    float(p.cost_price or 0),
        "selling_price": float(p.selling_price or 0),
        "supplier_id":   p.supplier_id,
        "supplier":      supplier,
        "created_at":    p.created_at,
    }


# ── CREATE ────────────────────────────────────────────────────────────────────
@router.post("/")
def create_product(
    product: schemas.ProductCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["admin", "manager"]))
):
    existing = db.query(models.Product).filter(
        models.Product.barcode     == product.barcode,
        models.Product.business_id == current_user.business_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Barcode already exists in your product catalog")

    supplier_id = getattr(product, "supplier_id", None)
    if supplier_id:
        supplier = db.query(models.Supplier).filter(
            models.Supplier.supplier_id == supplier_id,
            models.Supplier.business_id == current_user.business_id,
        ).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")

    new_product = models.Product(
        business_id=current_user.business_id,
        product_name=product.product_name,
        barcode=product.barcode,
        category_id=product.category_id,
        cost_price=product.cost_price,
        selling_price=product.selling_price,
        supplier_id=supplier_id,
    )
    db.add(new_product)
    db.flush()

    branches = _get_business_branches(db, current_user)
    for branch in branches:
        existing_inv = db.query(models.BranchInventory).filter(
            models.BranchInventory.product_id == new_product.product_id,
            models.BranchInventory.branch_id  == branch.branch_id
        ).first()
        if not existing_inv:
            db.add(models.BranchInventory(
                product_id=new_product.product_id,
                branch_id=branch.branch_id,
                stock_quantity=product.stock_quantity or 0,
                reorder_level=5,
            ))

    write_audit(
        db, current_user.user_id, "CREATE", "products", new_product.product_id,
        f"Added product '{new_product.product_name}' — barcode: {new_product.barcode} — price: ₦{float(new_product.selling_price):,.2f}",
    )

    db.commit()
    db.refresh(new_product)
    return _product_dict(new_product, db)


# ── LIST ──────────────────────────────────────────────────────────────────────
@router.get("/")
def get_products(
    search:      str = None,
    page:        int = 1,
    limit:       int = 20,
    supplier_id: int = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.Product)
    if current_user.role != SUPERADMIN_ROLE:
        query = query.filter(models.Product.business_id == current_user.business_id)
    if search:
        query = query.filter(models.Product.product_name.ilike(f"%{search}%"))
    if supplier_id:
        query = query.filter(models.Product.supplier_id == supplier_id)

    total    = query.count()
    products = query.offset((max(1, page) - 1) * limit).limit(limit).all()
    return {
        "total": total, "page": page, "limit": limit,
        "data":  [_product_dict(p, db) for p in products],
    }


# ── BARCODE LOOKUP ────────────────────────────────────────────────────────────
@router.get("/barcode/{barcode}")
def get_product_by_barcode(
    barcode: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.Product).filter(models.Product.barcode == barcode)
    if current_user.role != SUPERADMIN_ROLE:
        query = query.filter(models.Product.business_id == current_user.business_id)
    product = query.first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return _product_dict(product, db)


# ── GET ONE ───────────────────────────────────────────────────────────────────
@router.get("/{product_id}")
def get_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    product = db.query(models.Product).filter(
        models.Product.product_id == product_id
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if current_user.role != SUPERADMIN_ROLE and product.business_id != current_user.business_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    return _product_dict(product, db)


# ── UPDATE ────────────────────────────────────────────────────────────────────
@router.put("/{product_id}")
def update_product(
    product_id: int,
    product: schemas.ProductCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["admin", "manager"]))
):
    existing = db.query(models.Product).filter(
        models.Product.product_id == product_id
    ).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    if current_user.role != SUPERADMIN_ROLE and existing.business_id != current_user.business_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    changes = []
    if existing.product_name != product.product_name:
        changes.append(f"name: '{existing.product_name}' → '{product.product_name}'")
    if float(existing.selling_price or 0) != float(product.selling_price or 0):
        changes.append(f"price: ₦{float(existing.selling_price or 0):,.2f} → ₦{float(product.selling_price or 0):,.2f}")
    if float(existing.cost_price or 0) != float(product.cost_price or 0):
        changes.append(f"cost: ₦{float(existing.cost_price or 0):,.2f} → ₦{float(product.cost_price or 0):,.2f}")

    supplier_id = getattr(product, "supplier_id", None)
    if supplier_id and supplier_id != existing.supplier_id:
        supplier = db.query(models.Supplier).filter(
            models.Supplier.supplier_id == supplier_id,
            models.Supplier.business_id == current_user.business_id,
        ).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")
        changes.append("supplier changed")

    existing.product_name  = product.product_name
    existing.barcode       = product.barcode
    existing.category_id   = product.category_id
    existing.cost_price    = product.cost_price
    existing.selling_price = product.selling_price
    existing.supplier_id   = supplier_id

    write_audit(
        db, current_user.user_id, "UPDATE", "products", product_id,
        f"Updated '{existing.product_name}' (#{product_id}) — {', '.join(changes) if changes else 'no changes'}",
    )

    db.commit()
    db.refresh(existing)
    return _product_dict(existing, db)


# ── IMPORT TEMPLATE DOWNLOAD ──────────────────────────────────────────────────
@router.get("/import/template")
def download_import_template(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_role(["admin", "manager"]))
):
    """
    Returns a CSV template pre-populated with column headers and example rows.
    """
    from fastapi.responses import StreamingResponse

    rows = [
        "product_name,barcode,selling_price,cost_price,stock_quantity,category,supplier,expiry_date",
        '"Indomie Noodles (Chicken)",8712345678901,250,180,100,Food & Beverages,Dangote Suppliers,',
        '"Men Polo Shirt - Black",PT1234567890,4500,2800,20,Clothing,,',
        '"Paracetamol 500mg",6001234567890,150,80,50,Pharmaceuticals,Lagos Pharma Dist,2026-12-31',
    ]
    content = "\n".join(rows)
    return StreamingResponse(
        io.StringIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=profittrack_import_template.csv"}
    )


# ── IMPORT (xlsx or csv) ──────────────────────────────────────────────────────
@router.post("/import")
def import_products(
    file: UploadFile = File(...),
    db:   Session    = Depends(get_db),
    current_user: models.User = Depends(require_role(["admin", "manager"]))
):
    """
    Import/restock products from .xlsx or .csv file.

    Columns (header row required):
      product_name* | barcode* | selling_price* (new only) | cost_price |
      stock_quantity | category | supplier | expiry_date

    Rules:
    - NEW product (barcode not found)     → create product + inventory
    - EXISTING product (barcode matches)  → restock quantity only
    - Prices ignored for existing products (must be updated via Products page)
    - Supplier matched by name (case-insensitive) — if no match, product imported without supplier link
    - Unrecognised supplier names reported in warnings
    """
    filename = file.filename.lower()
    imported, restocked, skipped = 0, 0, 0
    errors   = []   # blocking row errors (row not processed)
    warnings = []   # non-blocking notices (row processed but something to note)

    # ── Helper: get or create category ───────────────────────────────────────
    def get_or_create_category(name: str):
        if not name: return None
        name = str(name).strip()
        cat = db.query(models.Category).filter(
            models.Category.category_name.ilike(name)
        ).first()
        if not cat:
            cat = models.Category(category_name=name)
            db.add(cat); db.flush()
        return cat.category_id

    # ── Helper: match supplier by name (this business only) ──────────────────
    def match_supplier(name: str):
        """
        Returns supplier_id if found, None if not.
        Never creates a new supplier — unmatched names are reported as warnings.
        """
        if not name: return None
        name = str(name).strip()
        supplier = db.query(models.Supplier).filter(
            models.Supplier.supplier_name.ilike(name),
            models.Supplier.business_id == current_user.business_id,
        ).first()
        return supplier.supplier_id if supplier else None

    branches = _get_business_branches(db, current_user)

    # ── Parse file ────────────────────────────────────────────────────────────
    try:
        if filename.endswith(".csv"):
            content = file.file.read().decode("utf-8-sig")
            reader  = csv.DictReader(io.StringIO(content))
            rows    = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            wb   = openpyxl.load_workbook(file.file, data_only=True)
            ws   = wb.active
            hdrs = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(hdrs, r)))
        else:
            raise HTTPException(status_code=400, detail="Only .csv or .xlsx files are supported")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    imported_names  = []
    restocked_names = []

    for i, row in enumerate(rows, start=2):
        # Normalise keys
        row = {str(k).strip().lower(): v for k, v in row.items()}

        # ── Extract fields ────────────────────────────────────────────────────
        product_name  = str(row.get("product_name") or "").strip()
        barcode       = str(row.get("barcode")       or "").strip()
        selling_price = row.get("selling_price")
        cost_price    = row.get("cost_price")
        stock_qty     = row.get("stock_quantity") or row.get("stock") or 0
        category_name = str(row.get("category")   or "").strip()
        supplier_name = str(row.get("supplier")    or "").strip()
        expiry_raw    = row.get("expiry_date")

        # ── Required field validation ─────────────────────────────────────────
        if not product_name:
            errors.append(f"Row {i}: missing product_name — row skipped"); continue
        if not barcode:
            errors.append(f"Row {i} ({product_name}): missing barcode — row skipped"); continue

        qty = 0
        if stock_qty:
            try:
                qty = int(float(str(stock_qty)))
                if qty < 0:
                    errors.append(f"Row {i} ({product_name}): quantity cannot be negative — row skipped"); continue
            except ValueError:
                errors.append(f"Row {i} ({product_name}): invalid quantity '{stock_qty}' — row skipped"); continue

        # ── Parse expiry date ─────────────────────────────────────────────────
        expiry_date = None
        if expiry_raw and str(expiry_raw).strip():
            try:
                raw = str(expiry_raw).strip()
                if isinstance(expiry_raw, str):
                    expiry_date = date.fromisoformat(raw)
                elif hasattr(expiry_raw, "date"):
                    expiry_date = expiry_raw.date()
                else:
                    expiry_date = date.fromisoformat(raw)
            except Exception:
                errors.append(f"Row {i} ({product_name}): invalid expiry_date '{expiry_raw}' — use YYYY-MM-DD — row skipped")
                continue

        # ── Supplier matching (warn but don't block) ──────────────────────────
        supplier_id = None
        if supplier_name:
            supplier_id = match_supplier(supplier_name)
            if supplier_id is None:
                warnings.append(
                    f"Row {i} ({product_name}): supplier '{supplier_name}' not found in your supplier list — "
                    f"product imported without supplier link. Add the supplier in the Suppliers page and edit the product to link it."
                )

        # ── Check if product exists in this business ──────────────────────────
        existing_product = db.query(models.Product).filter(
            models.Product.barcode     == barcode,
            models.Product.business_id == current_user.business_id,
        ).first()

        if existing_product:
            # ── RESTOCK: product exists ───────────────────────────────────────
            if qty > 0:
                for branch in branches:
                    inv = db.query(models.BranchInventory).filter(
                        models.BranchInventory.product_id == existing_product.product_id,
                        models.BranchInventory.branch_id  == branch.branch_id,
                    ).first()
                    if inv:
                        inv.stock_quantity += qty
                    else:
                        db.add(models.BranchInventory(
                            product_id=existing_product.product_id,
                            branch_id=branch.branch_id,
                            stock_quantity=qty,
                            reorder_level=5,
                        ))

                    db.add(models.InventoryMovement(
                        product_id=existing_product.product_id,
                        branch_id=branch.branch_id,
                        movement_type="RESTOCK",
                        reference_id=existing_product.product_id,
                        quantity=qty,
                        movement_date=now_lagos(),
                    ))

                    if expiry_date:
                        db.add(models.InventoryBatch(
                            product_id=existing_product.product_id,
                            branch_id=branch.branch_id,
                            quantity=qty,
                            expiry_date=expiry_date,
                            received_date=date.today(),
                            notes="Restocked via bulk upload",
                        ))

                # Update supplier link if not already set and we found a match
                if supplier_id and not existing_product.supplier_id:
                    existing_product.supplier_id = supplier_id

                restocked_names.append(f"{existing_product.product_name} (+{qty})")
                restocked += 1

            else:
                # Product exists but qty is 0 — nothing to do
                skipped += 1
            continue

        # ── NEW PRODUCT ───────────────────────────────────────────────────────
        # selling_price required for new products
        if not selling_price or str(selling_price).strip() == "":
            errors.append(
                f"Row {i} ({product_name}): missing selling_price — "
                f"required for new products — row skipped"
            )
            continue

        try:
            selling_price_f = float(str(selling_price).strip())
            if selling_price_f <= 0:
                errors.append(f"Row {i} ({product_name}): selling_price must be greater than 0 — row skipped")
                continue
        except ValueError:
            errors.append(f"Row {i} ({product_name}): invalid selling_price '{selling_price}' — row skipped")
            continue

        cost_price_f = 0.0
        if cost_price and str(cost_price).strip():
            try:
                cost_price_f = float(str(cost_price).strip())
            except ValueError:
                warnings.append(f"Row {i} ({product_name}): invalid cost_price '{cost_price}' — set to 0")

        try:
            category_id = get_or_create_category(category_name)

            product = models.Product(
                business_id=current_user.business_id,
                product_name=product_name,
                barcode=barcode,
                category_id=category_id,
                cost_price=cost_price_f,
                selling_price=selling_price_f,
                supplier_id=supplier_id,
            )
            db.add(product)
            db.flush()

            for branch in branches:
                db.add(models.BranchInventory(
                    product_id=product.product_id,
                    branch_id=branch.branch_id,
                    stock_quantity=qty,
                    reorder_level=5,
                    expiry_alert_days=90,
                ))
                if qty > 0 or expiry_date:
                    db.add(models.InventoryBatch(
                        product_id=product.product_id,
                        branch_id=branch.branch_id,
                        quantity=qty,
                        expiry_date=expiry_date,
                        received_date=date.today(),
                        notes="Imported via bulk upload",
                    ))

            imported_names.append(product_name)
            imported += 1

        except Exception as e:
            errors.append(f"Row {i} ({product_name}): unexpected error — {str(e)} — row skipped")
            db.rollback()
            continue

    # ── Audit log ─────────────────────────────────────────────────────────────
    if imported > 0 or restocked > 0:
        parts = []
        if imported > 0:
            preview = ", ".join(imported_names[:3])
            if len(imported_names) > 3: preview += f" +{len(imported_names)-3} more"
            parts.append(f"{imported} new: {preview}")
        if restocked > 0:
            preview = ", ".join(restocked_names[:3])
            if len(restocked_names) > 3: preview += f" +{len(restocked_names)-3} more"
            parts.append(f"{restocked} restocked: {preview}")
        write_audit(db, current_user.user_id, "CREATE", "products", 0,
                    f"Bulk upload — {'; '.join(parts)}")

    db.commit()

    return {
        "imported":  imported,
        "restocked": restocked,
        "skipped":   skipped,
        "errors":    errors,
        "warnings":  warnings,
        "message":   f"{imported} new products, {restocked} restocked, {skipped} skipped",
    }